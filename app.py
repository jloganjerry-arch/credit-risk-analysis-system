from flask import Flask, render_template, request, redirect, url_for, jsonify, session, make_response, flash
import pickle
import numpy as np
import pandas as pd
import sqlite3
import json
import shap
import warnings
import matplotlib
import os
import io
from datetime import datetime
from dotenv import load_dotenv
import groq
from groq import Groq
import bcrypt
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

matplotlib.use('Agg')
import matplotlib.pyplot as plt

warnings.filterwarnings('ignore')

# Load environment variables
load_dotenv()
# Note: google-genai client automatically picks up GEMINI_API_KEY from environment

# =========================================
# CREATE FLASK APP
# =========================================
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'credit-risk-secret-key-2025')

# =========================================
# FLASK-LOGIN SETUP
# =========================================
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'

class User(UserMixin):
    def __init__(self, id, username, role):
        self.id = id
        self.username = username
        self.role = role

@login_manager.user_loader
def load_user(user_id):
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, role FROM users WHERE id=?', (user_id,))
    row = cursor.fetchone()
    conn.close()
    if row:
        return User(row[0], row[1], row[2])
    return None

# =========================================
# CREATE DATABASE
# =========================================
conn = sqlite3.connect('history.db', check_same_thread=False)
cursor = conn.cursor()
cursor.execute('''
    CREATE TABLE IF NOT EXISTS predictions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        contract_type TEXT,
        credit_amount REAL,
        age REAL,
        result TEXT,
        risk_score REAL,
        occupation_type TEXT,
        risk_reason TEXT
    )
''')
# Users table for multi-user auth
cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'analyst'
    )
''')
# Alter table to add risk_reason if it doesn't exist from older versions
try:
    cursor.execute('ALTER TABLE predictions ADD COLUMN risk_reason TEXT')
except:
    pass

# Seed default admin user if no users exist
cursor.execute('SELECT COUNT(*) FROM users')
if cursor.fetchone()[0] == 0:
    hashed = bcrypt.hashpw('admin123'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    cursor.execute('INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                   ('admin', hashed, 'admin'))
conn.commit()

# =========================================
# LOAD TRAINED MODEL & PROCESSORS
# =========================================
try:
    model = pickle.load(open('credit_fraud_model.pkl', 'rb'))
    label_encoders = pickle.load(open('label_encoders.pkl', 'rb'))
    imputer = pickle.load(open('imputer.pkl', 'rb'))
    explainer = pickle.load(open('shap_explainer.pkl', 'rb'))
except Exception as e:
    print(f"Warning: Model files missing or corrupt. Please run train_model.py first. Error: {e}")
    model, label_encoders, imputer, explainer = None, None, None, None

features = [
    'NAME_CONTRACT_TYPE',
    'DAYS_BIRTH',
    'AMT_CREDIT',
    'AMT_INCOME_TOTAL',
    'NAME_INCOME_TYPE',
    'NAME_EDUCATION_TYPE',
    'OCCUPATION_TYPE',
    'DAYS_EMPLOYED'
]

# =========================================
# AI ASSISTANT CHATBOT PAGE
# =========================================
@app.route('/chat')
@login_required
def chat():
    return render_template('chat.html')

# =========================================
# API: AI CHAT (with Conversation Memory)
# =========================================
@app.route('/api/chat', methods=['POST'])
def api_chat():
    data = request.get_json()
    user_message = data.get('message', '')

    system_prompt = """You are a specialized Credit Risk Analysis Assistant.
You MUST ONLY answer questions related to:
- Credit Risk Analysis
- Risk Scores
- Prediction Results
- SHAP Explanations
- Model Features
- Credit Amount, Income, Occupation, Education, Employment Analysis
- Project Functionality

If a user asks an unrelated question (e.g. "Who won the IPL?", "Write a poem", "What is the capital of France?", etc.), you MUST decline by responding exactly with:
"I am a specialized Credit Risk Analysis Assistant and can only answer questions related to transaction risk assessment and prediction results."

Maintain a professional, banking-style tone at all times.
"""
    # Load existing conversation history from session
    if 'chat_history' not in session:
        session['chat_history'] = []

    # Append new user message
    session['chat_history'].append({"role": "user", "content": user_message})

    # Build full message list with system prompt
    messages = [{"role": "system", "content": system_prompt}] + session['chat_history']

    try:
        client = Groq(api_key=os.environ.get("GROQ_API_KEY"))
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=messages
        )
        reply = response.choices[0].message.content
        # Save AI reply to history
        session['chat_history'].append({"role": "assistant", "content": reply})
        session.modified = True
        return jsonify({"reply": reply})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =========================================
# API: CLEAR CHAT HISTORY
# =========================================
@app.route('/api/chat/clear', methods=['POST'])
def clear_chat():
    session.pop('chat_history', None)
    return jsonify({"status": "cleared"})

# =========================================
# API: WHAT-IF SCENARIO SIMULATOR
# =========================================
@app.route('/api/simulate', methods=['POST'])
def api_simulate():
    if model is None:
        return jsonify({"error": "Model not loaded"}), 500
    try:
        data = request.get_json()
        raw_data = {
            'NAME_CONTRACT_TYPE': data.get('contract_type', 'Cash loans'),
            'DAYS_BIRTH': float(data.get('age', 30)) * 365,
            'AMT_CREDIT': float(data.get('credit', 300000)),
            'AMT_INCOME_TOTAL': float(data.get('income', 150000)),
            'NAME_INCOME_TYPE': data.get('income_type', 'Working'),
            'NAME_EDUCATION_TYPE': data.get('education', 'Higher education'),
            'OCCUPATION_TYPE': data.get('occupation_type', 'Managers'),
            'DAYS_EMPLOYED': float(data.get('days_employed', 365))
        }
        age = float(data.get('age', 30))
        credit = raw_data['AMT_CREDIT']
        income = raw_data['AMT_INCOME_TOTAL']
        days_employed = raw_data['DAYS_EMPLOYED']

        input_df = pd.DataFrame([raw_data])
        input_df_imputed = pd.DataFrame(imputer.transform(input_df), columns=features)
        input_df_imputed['DAYS_BIRTH'] = abs(input_df_imputed['DAYS_BIRTH'].astype(float)) / 365
        input_df_imputed['DAYS_EMPLOYED'] = abs(input_df_imputed['DAYS_EMPLOYED'].astype(float))
        for col in ['NAME_CONTRACT_TYPE', 'NAME_INCOME_TYPE', 'NAME_EDUCATION_TYPE', 'OCCUPATION_TYPE']:
            val = str(input_df_imputed[col].iloc[0])
            if val in label_encoders[col].classes_:
                input_df_imputed[col] = label_encoders[col].transform([val])[0]
            else:
                input_df_imputed[col] = 0
        for col in input_df_imputed.columns:
            input_df_imputed[col] = input_df_imputed[col].astype(float)

        prediction = model.predict(input_df_imputed)
        if hasattr(model, 'predict_proba'):
            probabilities = model.predict_proba(input_df_imputed)
            risk_score = round(probabilities[0][1] * 100, 2)
        else:
            risk_score = 100.0 if prediction[0] == 1 else 0.0

        # Apply rule-based overrides
        if age < 25 and income < 30000 and credit > 250000:
            risk_score = max(risk_score, 98.0)
        elif days_employed < 180 and credit > 200000:
            risk_score = max(risk_score, 95.0)
        elif income > 0 and (credit / income) > 10:
            risk_score = max(risk_score, 96.0)

        if risk_score >= 70:
            risk_category = "High Risk"
        elif risk_score >= 40:
            risk_category = "Medium Risk"
        else:
            risk_category = "Low Risk"

        return jsonify({"risk_score": risk_score, "risk_category": risk_category})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =========================================
# API: ANALYZE WITH AI
# =========================================
@app.route('/api/analyze', methods=['POST'])
def api_analyze():
    data = request.get_json()
    
    prompt = f"""You are an expert AI Credit Risk Analyst. Analyze the following prediction data and generate a structured risk assessment report.

Prediction Data:
- Risk Category: {data.get('risk_category')}
- Risk Score: {data.get('risk_score')}%
- Contract Type: {data.get('contract_type')}
- Credit Amount: {data.get('credit')}
- Total Income: {data.get('income')}
- Age Profile: {data.get('age')} years
- Employment History: {data.get('days_employed')} days
- Occupation: {data.get('occupation')}
- Education: {data.get('education')}
- SHAP/Rule Insights: {data.get('risk_reason')}

Generate a structured output exactly containing these sections (use Markdown):
### AI Risk Assessment
[Provide a clear summary sentence like "This applicant was classified as <Category> because..."]

### Key Risk Factors
[Bullet points of the main factors driving this score]

### SHAP Insights
[Explain how the AI weighted the specific features (SHAP values)]

### Potential Concerns
[Any red flags or specific warnings]

### Recommendation
[Actionable recommendation for the loan officer (e.g., approve, manual verification, reject)]

Maintain a highly professional, banking-style tone.
"""
    try:
        client = Groq(api_key=os.environ.get("GROQ_API_KEY"))
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": "You are an expert AI Credit Risk Analyst."},
                {"role": "user", "content": prompt}
            ]
        )
        return jsonify({"analysis": response.choices[0].message.content})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =========================================
# LOGIN PAGE
# =========================================
@app.route('/')
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def validate_login():
    username = request.form['username']
    password = request.form['password']
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, password_hash, role FROM users WHERE username=?', (username,))
    row = cursor.fetchone()
    conn.close()
    if row and bcrypt.checkpw(password.encode('utf-8'), row[2].encode('utf-8')):
        user = User(row[0], row[1], row[3])
        login_user(user)
        return redirect(url_for('home'))
    return render_template('login.html', error='Invalid username or password.')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# =========================================
# REGISTER PAGE (Admin Only)
# =========================================
@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        role = request.form['role']
        if not username or not password:
            return render_template('register.html', error='All fields are required.')
        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        try:
            conn = sqlite3.connect('history.db')
            cursor = conn.cursor()
            cursor.execute('INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                           (username, hashed, role))
            conn.commit()
            conn.close()
            return render_template('register.html', success=f'User "{username}" created successfully!')
        except sqlite3.IntegrityError:
            return render_template('register.html', error=f'Username "{username}" already exists.')
    return render_template('register.html')

# =========================================
# HOME PAGE
# =========================================
@app.route('/home')
@login_required
def home():
    return render_template('index.html', username=current_user.username, role=current_user.role)

# =========================================
# PREDICTION ROUTE
# =========================================
@app.route('/predict', methods=['POST'])
@login_required
def predict():
    if model is None:
        return render_template('index.html', prediction_text="Model not trained yet. Please run training script.")

    # Get raw inputs
    raw_data = {
        'NAME_CONTRACT_TYPE': request.form['contract_type'],
        'DAYS_BIRTH': float(request.form['age']) * 365,  # Convert back to raw DAYS for imputer match
        'AMT_CREDIT': float(request.form['credit']),
        'AMT_INCOME_TOTAL': float(request.form['income']),
        'NAME_INCOME_TYPE': request.form['income_type'],
        'NAME_EDUCATION_TYPE': request.form['education'],
        'OCCUPATION_TYPE': request.form['occupation_type'],
        'DAYS_EMPLOYED': float(request.form['days_employed'])
    }
    
    age = float(request.form['age'])
    credit = raw_data['AMT_CREDIT']
    income = raw_data['AMT_INCOME_TOTAL']
    days_employed = raw_data['DAYS_EMPLOYED']

    # Validation
    if age < 18 or age > 100:
        return render_template('index.html', prediction_text="Invalid age entered")

    # Dynamic Preprocessing matching training script
    input_df = pd.DataFrame([raw_data])
    
    # Impute
    input_df_imputed = pd.DataFrame(imputer.transform(input_df), columns=features)
    
    # Feature Engineering
    input_df_imputed['DAYS_BIRTH'] = abs(input_df_imputed['DAYS_BIRTH'].astype(float)) / 365
    input_df_imputed['DAYS_EMPLOYED'] = abs(input_df_imputed['DAYS_EMPLOYED'].astype(float))
    
    # Encode categorical
    for col in ['NAME_CONTRACT_TYPE', 'NAME_INCOME_TYPE', 'NAME_EDUCATION_TYPE', 'OCCUPATION_TYPE']:
        val = str(input_df_imputed[col].iloc[0])
        # Handle unseen labels gracefully by assigning to a known class (or 0)
        if val in label_encoders[col].classes_:
            input_df_imputed[col] = label_encoders[col].transform([val])[0]
        else:
            input_df_imputed[col] = 0
            
    # Convert all to float
    for col in input_df_imputed.columns:
        input_df_imputed[col] = input_df_imputed[col].astype(float)

    # ML Model Prediction
    prediction = model.predict(input_df_imputed)
    if hasattr(model, 'predict_proba'):
        probabilities = model.predict_proba(input_df_imputed)
        risk_score = round(probabilities[0][1] * 100, 2)
    else:
        risk_score = 100.0 if prediction[0] == 1 else 0.0
        
    is_high_risk = (prediction[0] == 1)

    # =========================================
    # RULE-BASED VALIDATION OVERRIDES
    # =========================================
    rule_reason = ""
    if age < 25 and income < 30000 and credit > 250000:
        is_high_risk = True
        risk_score = max(risk_score, 98.0)
        rule_reason = "Flagged due to young age, low income, and excessively high requested credit."
    elif str(raw_data['NAME_INCOME_TYPE']).lower() == 'student' and credit > 100000:
        is_high_risk = True
        risk_score = max(risk_score, 99.0)
        rule_reason = "Flagged due to unusually high loan request for a student."
    elif days_employed < 180 and credit > 200000:
        is_high_risk = True
        risk_score = max(risk_score, 95.0)
        rule_reason = "Flagged due to very low employment history combined with high credit requirement."
    elif income > 0 and (credit / income) > 10:
        is_high_risk = True
        risk_score = max(risk_score, 96.0)
        rule_reason = "Flagged due to dangerously high credit-to-income ratio (>10x)."

    # =========================================
    # RISK CATEGORY MAPPING (3-Tier)
    # =========================================
    if risk_score >= 70:
        risk_category = "High Risk"
    elif risk_score >= 40:
        risk_category = "Medium Risk"
    else:
        risk_category = "Low Risk"

    # Rule-based overrides override category
    if rule_reason:
        risk_category = "High Risk"

    # =========================================
    # SHAP HUMAN-READABLE EXPLANATION
    # =========================================
    shap_vals = explainer.shap_values(input_df_imputed)
    if isinstance(shap_vals, list):
        shap_vals = shap_vals[1]
    
    base_value = explainer.expected_value
    if isinstance(base_value, list):
        base_value = base_value[1]
    elif isinstance(base_value, np.ndarray) and len(base_value) > 1:
        base_value = base_value[1]
        
    explanation = shap.Explanation(
        values=shap_vals[0],
        base_values=base_value,
        data=input_df_imputed.iloc[0].values,
        feature_names=features
    )
    
    # Generate and save Waterfall Plot
    plt.figure(figsize=(7, 4))
    shap.plots.waterfall(explanation, show=False)
    plt.tight_layout()
    plt.savefig('static/local_shap.png', bbox_inches='tight', dpi=150)
    plt.close()

    # Sort features by absolute impact for textual reason
    impacts = shap_vals[0]
    feature_impacts = [(features[i], impacts[i]) for i in range(len(features))]
    feature_impacts.sort(key=lambda x: abs(x[1]), reverse=True)
    
    top_pushing_risk = [f[0] for f in feature_impacts if f[1] > 0][:3]
    top_lowering_risk = [f[0] for f in feature_impacts if f[1] < 0][:3]
    
    # Map raw features to human names and explanations
    feature_name_map = {
        'NAME_CONTRACT_TYPE': 'Contract type',
        'DAYS_BIRTH': 'Age profile',
        'AMT_CREDIT': 'Credit amount',
        'AMT_INCOME_TOTAL': 'Total income',
        'NAME_INCOME_TYPE': 'Income source',
        'NAME_EDUCATION_TYPE': 'Education level',
        'OCCUPATION_TYPE': 'Occupation',
        'DAYS_EMPLOYED': 'Employment history'
    }
    
    risk_reason_html = "<strong>💡 AI Explanation</strong><br><br>"
    
    if rule_reason:
        risk_reason_html += f"This application was classified as HIGH RISK due to a critical rule violation:<br><br><ul><li>{rule_reason}</li></ul>"
        risk_reason_html += "<br><strong>Recommendation:</strong><br>Immediate rejection or manual verification required."
    else:
        if risk_category == "High Risk":
            risk_reason_html += "This application was classified as HIGH RISK because:<br><br><ul>"
            for f in top_pushing_risk:
                risk_reason_html += f"<li>{feature_name_map.get(f, f)} contributes negatively to repayment confidence.</li>"
            risk_reason_html += "</ul><br><strong>Recommendation:</strong><br>Manual verification is advised before approval."
            
        elif risk_category == "Medium Risk":
            risk_reason_html += "This application was classified as MEDIUM RISK. Key factors include:<br><br><ul>"
            if top_pushing_risk:
                risk_reason_html += f"<li>{feature_name_map.get(top_pushing_risk[0], top_pushing_risk[0])} requires attention.</li>"
            if top_lowering_risk:
                risk_reason_html += f"<li>{feature_name_map.get(top_lowering_risk[0], top_lowering_risk[0])} provides some stability.</li>"
            risk_reason_html += "</ul><br><strong>Recommendation:</strong><br>Proceed with caution and standard verification."
            
        else:
            risk_reason_html += "This application was classified as LOW RISK because:<br><br><ul>"
            for f in top_lowering_risk:
                risk_reason_html += f"<li>{feature_name_map.get(f, f)} strongly supports repayment capability.</li>"
            risk_reason_html += "</ul><br><strong>Recommendation:</strong><br>Standard processing approved."

    # =========================================
    # SAVE TO DATABASE
    # =========================================
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO predictions (contract_type, credit_amount, age, result, risk_score, occupation_type, risk_reason)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (str(raw_data['NAME_CONTRACT_TYPE']), float(credit), float(age), risk_category, float(risk_score), str(raw_data['OCCUPATION_TYPE']), risk_reason_html))
    tx_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return render_template('index.html', prediction_text=risk_category, risk_score=risk_score, risk_reason=risk_reason_html, show_results=True, transaction_id=tx_id)

# =========================================
# HISTORY PAGE
# =========================================
@app.route('/history')
@login_required
def history():
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM predictions ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()
    return render_template('history.html', predictions=rows, username=current_user.username, role=current_user.role)

# =========================================
# ANALYSIS PAGE
# =========================================
@app.route('/analysis')
@login_required
def analysis():
    if current_user.role not in ['admin', 'manager']:
        return redirect(url_for('home'))
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute("SELECT result, COUNT(*) FROM predictions GROUP BY result")
    results = cursor.fetchall()
    high_risk_count = 0
    low_risk_count = 0
    for res in results:
        if 'High Risk' in res[0]:
            high_risk_count += res[1]
        else:
            low_risk_count += res[1]

    cursor.execute("SELECT credit_amount FROM predictions")
    credit_amounts = [row[0] for row in cursor.fetchall()]
    conn.close()

    try:
        with open('model_metrics.json', 'r') as f:
            metrics = json.load(f)
    except FileNotFoundError:
        metrics = None

    return render_template('analysis.html', high_risk=high_risk_count, low_risk=low_risk_count, credit_amounts=credit_amounts, metrics=metrics)

# =========================================
# EXPORT PDF REPORT
# =========================================
@app.route('/export_pdf/<int:transaction_id>')
def export_pdf(transaction_id):
    # Fetch transaction from DB
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM predictions WHERE id=?", (transaction_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return "Transaction not found", 404

    tx_id, contract_type, credit_amount, age, risk_category, risk_score, occupation_type, risk_reason = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7] if len(row) > 7 else ''

    # Generate AI narrative via Groq
    ai_narrative = "AI analysis not available (API key not configured)."
    try:
        client = Groq(api_key=os.environ.get("GROQ_API_KEY"))
        prompt = f"""You are a senior credit risk analyst writing an official bank report.
Write a concise 3-paragraph professional risk assessment for this loan application:
- Transaction ID: #{tx_id}
- Contract Type: {contract_type}
- Credit Amount: Rs {credit_amount:,.2f}
- Applicant Age: {age} years
- Occupation: {occupation_type}
- Risk Score: {risk_score}%
- Risk Category: {risk_category}

Paragraph 1: Executive Summary (2-3 sentences summarizing the decision)
Paragraph 2: Key Risk Factors (what drove this risk score)
Paragraph 3: Recommendation (specific actionable advice for loan officer)

Use formal banking language. Do NOT use markdown or asterisks."""
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}]
        )
        ai_narrative = response.choices[0].message.content
    except Exception as e:
        ai_narrative = f"AI narrative unavailable: {str(e)}"

    # Build PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Color scheme
    primary_color = colors.HexColor('#1e3a8a')
    danger_color  = colors.HexColor('#ef4444')
    success_color = colors.HexColor('#10b981')
    warning_color = colors.HexColor('#f59e0b')
    risk_color = danger_color if 'High' in str(risk_category) else (warning_color if 'Medium' in str(risk_category) else success_color)

    # Title Header
    title_style = ParagraphStyle('Title', parent=styles['Title'],
        fontSize=20, textColor=primary_color, spaceAfter=4, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#64748b'), alignment=TA_CENTER, spaceAfter=2)
    story.append(Paragraph("CREDIT RISK ASSESSMENT REPORT", title_style))
    story.append(Paragraph("Transaction Risk Analyzer — Powered by AI", subtitle_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}", subtitle_style))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width='100%', thickness=2, color=primary_color))
    story.append(Spacer(1, 0.4*cm))

    # Risk Score Banner
    score_label = str(risk_category).upper() if risk_category else 'UNKNOWN'
    score_val   = f"{float(risk_score):.1f}%" if risk_score is not None else 'N/A'
    banner_data = [[f'RISK CATEGORY: {score_label}', f'RISK SCORE: {score_val}']]
    banner_table = Table(banner_data, colWidths=[9*cm, 8*cm])
    banner_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), risk_color),
        ('TEXTCOLOR',  (0,0), (-1,-1), colors.white),
        ('FONTNAME',   (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 13),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [risk_color]),
        ('TOPPADDING',    (0,0), (-1,-1), 12),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ('ROUNDEDCORNERS', [6]),
    ]))
    story.append(banner_table)
    story.append(Spacer(1, 0.5*cm))

    # Applicant Details Table
    section_style = ParagraphStyle('Section', parent=styles['Heading2'],
        fontSize=12, textColor=primary_color, spaceBefore=8, spaceAfter=4)
    story.append(Paragraph("Applicant Details", section_style))
    details_data = [
        ['Field', 'Value'],
        ['Transaction ID', f'#{tx_id}'],
        ['Contract Type', str(contract_type).title()],
        ['Credit Amount', f'Rs {float(credit_amount):,.2f}'],
        ['Applicant Age', f'{float(age):.0f} years'],
        ['Occupation', str(occupation_type).title()],
    ]
    details_table = Table(details_data, colWidths=[6*cm, 11*cm])
    details_table.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), primary_color),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,-1), 10),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8fafc'), colors.white]),
        ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING',  (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(details_table)
    story.append(Spacer(1, 0.5*cm))

    # AI Narrative
    story.append(Paragraph("AI Risk Assessment Narrative", section_style))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#cbd5e1')))
    story.append(Spacer(1, 0.2*cm))
    body_style = ParagraphStyle('Body', parent=styles['Normal'],
        fontSize=10, leading=16, textColor=colors.HexColor('#1e293b'), spaceAfter=8)
    for para in ai_narrative.strip().split('\n'):
        if para.strip():
            story.append(Paragraph(para.strip(), body_style))
    story.append(Spacer(1, 0.5*cm))

    # Footer
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#cbd5e1')))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'],
        fontSize=8, textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER, spaceBefore=4)
    story.append(Paragraph("This report was automatically generated by the Transaction Risk Analyzer. It is intended for internal use by authorized loan officers only.", footer_style))
    story.append(Paragraph("© 2025 Risk Analysis System — Confidential", footer_style))

    doc.build(story)
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=risk_report_{tx_id}.pdf'
    return response

# =========================================
# BATCH PREDICT (CSV Upload)
# =========================================
# =========================================
# DOWNLOAD TEMPLATE CSV
# =========================================
@app.route('/download_template')
@login_required
def download_template():
    csv_content = (
        "contract_type,age,credit,income,income_type,education,occupation_type,days_employed\n"
        "Cash loans,42,450000,180000,Working,Secondary / special education,Laborers,1200\n"
        "Revolving loans,29,200000,120000,Commercial associate,Higher education,Core staff,450\n"
    )
    response = make_response(csv_content)
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=applicant_batch_template.csv'
    return response

# =========================================
# BATCH PREDICT (CSV Upload with Validation)
# =========================================
@app.route('/batch_predict', methods=['GET', 'POST'])
@login_required
def batch_predict():
    if request.method == 'GET':
        return render_template('batch_results.html',
                               username=current_user.username, role=current_user.role,
                               results=None)
    if model is None:
        return render_template('batch_results.html', error='Model not loaded.',
                               username=current_user.username, role=current_user.role, results=None)
    if 'csv_file' not in request.files or request.files['csv_file'].filename == '':
        return render_template('batch_results.html', error='No file selected.',
                               username=current_user.username, role=current_user.role, results=None)

    file = request.files['csv_file']
    try:
        df = pd.read_csv(file)
    except Exception as e:
        return render_template('batch_results.html', error=f'Could not read CSV: {e}',
                               username=current_user.username, role=current_user.role, results=None)

    # Normalize column names
    df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]

    # Required columns check
    required_cols = ['contract_type', 'age', 'credit', 'income', 'income_type', 'education', 'occupation_type', 'days_employed']
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        return render_template('batch_results.html', error=f"Missing required columns in CSV: {', '.join(missing_cols)}",
                               username=current_user.username, role=current_user.role, results=None)

    results = []
    db_conn = sqlite3.connect('history.db')
    db_cursor = db_conn.cursor()

    for idx, row in df.iterrows():
        try:
            errors = []
            
            raw_contract = row.get('contract_type')
            if pd.isna(raw_contract) or str(raw_contract).strip() == '':
                errors.append("contract_type is empty")
            else:
                raw_contract = str(raw_contract).strip()

            raw_age = row.get('age')
            try:
                age_val = float(raw_age)
                if pd.isna(raw_age) or age_val <= 0 or age_val > 120:
                    errors.append(f"age must be between 1 and 120 (got '{raw_age}')")
            except (ValueError, TypeError):
                errors.append(f"age must be a number (got '{raw_age}')")

            raw_credit = row.get('credit')
            try:
                credit_val = float(raw_credit)
                if pd.isna(raw_credit) or credit_val < 0:
                    errors.append(f"credit must be a positive number (got '{raw_credit}')")
            except (ValueError, TypeError):
                errors.append(f"credit must be a number (got '{raw_credit}')")

            raw_income = row.get('income')
            try:
                income_val = float(raw_income)
                if pd.isna(raw_income) or income_val < 0:
                    errors.append(f"income must be a positive number (got '{raw_income}')")
            except (ValueError, TypeError):
                errors.append(f"income must be a number (got '{raw_income}')")

            raw_income_type = row.get('income_type')
            if pd.isna(raw_income_type) or str(raw_income_type).strip() == '':
                errors.append("income_type is empty")

            raw_education = row.get('education')
            if pd.isna(raw_education) or str(raw_education).strip() == '':
                errors.append("education is empty")

            raw_occupation_type = row.get('occupation_type')
            if pd.isna(raw_occupation_type) or str(raw_occupation_type).strip() == '':
                errors.append("occupation_type is empty")

            raw_days_employed = row.get('days_employed')
            try:
                days_employed_val = float(raw_days_employed)
                if pd.isna(raw_days_employed):
                    errors.append("days_employed must be a valid number")
            except (ValueError, TypeError):
                errors.append(f"days_employed must be a number (got '{raw_days_employed}')")

            if errors:
                raise ValueError("; ".join(errors))

            raw_data = {
                'NAME_CONTRACT_TYPE': str(raw_contract),
                'DAYS_BIRTH': age_val * 365,
                'AMT_CREDIT': credit_val,
                'AMT_INCOME_TOTAL': income_val,
                'NAME_INCOME_TYPE': str(raw_income_type).strip(),
                'NAME_EDUCATION_TYPE': str(raw_education).strip(),
                'OCCUPATION_TYPE': str(raw_occupation_type).strip(),
                'DAYS_EMPLOYED': days_employed_val
            }
            age = age_val
            credit = credit_val
            income = income_val
            days_employed = days_employed_val

            input_df = pd.DataFrame([raw_data])
            input_df_imputed = pd.DataFrame(imputer.transform(input_df), columns=features)
            input_df_imputed['DAYS_BIRTH'] = abs(input_df_imputed['DAYS_BIRTH'].astype(float)) / 365
            input_df_imputed['DAYS_EMPLOYED'] = abs(input_df_imputed['DAYS_EMPLOYED'].astype(float))
            for col in ['NAME_CONTRACT_TYPE', 'NAME_INCOME_TYPE', 'NAME_EDUCATION_TYPE', 'OCCUPATION_TYPE']:
                val = str(input_df_imputed[col].iloc[0])
                input_df_imputed[col] = label_encoders[col].transform([val])[0] if val in label_encoders[col].classes_ else 0
            for col in input_df_imputed.columns:
                input_df_imputed[col] = input_df_imputed[col].astype(float)

            proba = model.predict_proba(input_df_imputed)[0][1] if hasattr(model, 'predict_proba') else 0.5
            risk_score = round(proba * 100, 2)

            # Rule overrides
            if age < 25 and income < 30000 and credit > 250000: risk_score = max(risk_score, 98.0)
            elif days_employed < 180 and credit > 200000: risk_score = max(risk_score, 95.0)
            elif income > 0 and (credit / income) > 10: risk_score = max(risk_score, 96.0)

            risk_category = 'High Risk' if risk_score >= 70 else ('Medium Risk' if risk_score >= 40 else 'Low Risk')

            db_cursor.execute("""
                INSERT INTO predictions (contract_type, credit_amount, age, result, risk_score, occupation_type, risk_reason)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (raw_data['NAME_CONTRACT_TYPE'], credit, age, risk_category, risk_score,
                  raw_data['OCCUPATION_TYPE'], 'Batch prediction'))

            results.append({
                'row': idx + 1,
                'contract_type': raw_data['NAME_CONTRACT_TYPE'],
                'age': age,
                'credit': credit,
                'income': income,
                'occupation': raw_data['OCCUPATION_TYPE'],
                'risk_score': risk_score,
                'risk_category': risk_category
            })
        except Exception as e:
            results.append({'row': idx + 1, 'error': str(e)})

    db_conn.commit()
    db_conn.close()
    return render_template('batch_results.html', results=results,
                           username=current_user.username, role=current_user.role,
                           total=len(results),
                           high=sum(1 for r in results if r.get('risk_category') == 'High Risk'),
                           medium=sum(1 for r in results if r.get('risk_category') == 'Medium Risk'),
                           low=sum(1 for r in results if r.get('risk_category') == 'Low Risk'))

# =========================================
# EXPORT EXCEL
# =========================================
@app.route('/export_excel')
@login_required
def export_excel():
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM predictions ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transaction History'

    # Styles
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='1E3A8A')
    high_fill    = PatternFill('solid', fgColor='FEE2E2')
    medium_fill  = PatternFill('solid', fgColor='FEF3C7')
    low_fill     = PatternFill('solid', fgColor='D1FAE5')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border  = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    headers = ['ID', 'Contract Type', 'Credit Amount (Rs)', 'Age', 'Risk Category', 'Risk Score (%)', 'Occupation']
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row in rows:
        risk_cat = str(row[4]) if row[4] else ''
        fill = high_fill if 'High' in risk_cat else (medium_fill if 'Medium' in risk_cat else low_fill)
        ws.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6]])
        for col_num in range(1, 8):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

    # Auto-fit columns
    col_widths = [6, 18, 20, 8, 16, 15, 16]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=transaction_history.xlsx'
    return response

# =========================================
# RUN FLASK APP
# =========================================
if __name__ == "__main__":
    app.run(debug=True)